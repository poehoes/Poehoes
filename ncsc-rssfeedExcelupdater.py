# parser that imports the ncsc-excelsheet. This file should be sorted at column 1 (ncsc-id)
# it will than download the rss-feed from the ncsc, parses all the ncsc-advisories and
# 1) add them to the bottom of the excelsheet
# 2) update the Version-column if the ncsc-advisory has been updated.

import feedparser
import os
import openpyxl
#from openpyxl.styles import colors
from openpyxl.styles import Font

cwd = os.getcwd()
# print the current directory 
print("Current working directory is:", cwd)

#use this script online with ncsc-feed or offline for rss importing xml/html export
online = True
ncscrssurl = 'https://advisories.ncsc.nl/rss/advisories'
xmlimport_filename = 'ncscexport20210129-01.xml'

#the excelsheet's filename
#ncsc_sheet_filename = "G:\\My Drive\\Security and Compliance\\5. IB\\Beheer\\Periodic_controls\\NCSC\\ncscnew.xlsx"
ncsc_sheet_filename = 'ncscnew.xlsx'

# No editing needed below this line
# --------------------------------

#scrape rss-feed from NCSC or import rss thru local xml-file
if online:
    try:
        #from url, enable if online
        scraperss =  feedparser.parse(ncscrssurl)
    except:
        print ("Url {} not available!".format(ncscrssurl))
else:
    #from a file, enable if local feed
    try:
        scraperss =  feedparser.parse(xmlimport_filename)
    except:
        print ("XML-file {} not found!".format(xmlimport_filename))

# open excelsheet
wb = openpyxl.load_workbook(ncsc_sheet_filename)
activeSheet = wb['Sheet1']
rssData = {}
sheetData = {}

#print("Reading rows...")
for row in range(2, activeSheet.max_row + 1):
    ncsc    = activeSheet['A' + str(row)].value
    version = activeSheet['B' + str(row)].value
    kansniveau    = activeSheet['C' + str(row)].value
    schadeniveau  = activeSheet['D' + str(row)].value
    pubdat  = activeSheet['E' + str(row)].value
    product = activeSheet['F' + str(row)].value
    beoorIB = activeSheet['G' + str(row)].value
    beoorDate     = activeSheet['H' + str(row)].value
    vuln    = activeSheet['I' + str(row)].value
    uri     = activeSheet['J' + str(row)].value
    sheetData.setdefault(ncsc,  {'Version':version, 'Vnew':'no', 'kansniveau':kansniveau, 'schadeniveau':schadeniveau, 'publicatie datum':pubdat, 'product':product, 'Beoordeling IB':beoorIB, 'Beoordeeld':beoorDate, 'Kwetsbaar':vuln, 'uri':uri})

lastInc = activeSheet['A2']
lastIncVers = activeSheet['B2']
#print("Previous message found, last message in excelsheet is: " + str(lastInc.value) + " Version " + str(lastIncVers.value) + "\n\n")

for post in scraperss.entries:
    #print(post.title + ": " + post.link)
    titel = post.title
    ncscid = titel[0:14]
    ncscvers = titel[16:20]
    start = titel.find('] [') + 3
    kans = titel[start]
    schade = titel[start+2]
    productline = titel[28:]
    #should we clean the productline by deletion of words like 'kwetsbaarheid, kwetsbaarheden, gevonden, verholpen, meerdere, verhelpt, ontdekt, producten
    if online:
        try:
            ncscdate = str(post.published_parsed[2]) + "-" + str(post.published_parsed[1]) + "-" + str(post.published_parsed[0])
        except:
            print ("Date online not correct!")
    else:
        try:
            ncscdate = str(post.published)
        except:
            print("Date in file not correct")
    ncscurl = post.link
    # determine kansniveau
    if kans == 'L':
        kansniveau = "laag"
    elif kans == 'M':
        kansniveau = "gemiddeld"
    elif kans == 'H':
        kansniveau = 'hoog'
    # determine schadeniveau    
    if schade == 'L':
        schadeniveau = "laag"
    elif schade == 'M':
        schadeniveau = "gemiddeld"
    elif schade == 'H':
        schadeniveau = 'hoog'
    if schade == 'H' and kans == 'H':
        print(ncscid, "Kansniveau:" + kansniveau,"Schadeniveau:" + schadeniveau, productline)
    rssData.setdefault(ncscid, {'Version':ncscvers, 'Vnew':'no', 'kansniveau':kansniveau, 'schadeniveau':schadeniveau, 'publicatie datum':ncscdate, 'product':productline, 'Beoordeling IB':'', 'Beoordeeld':'', 'Kwetsbaar':'', 'uri':ncscurl})
    #print(ncscid, ncscvers,"Kansniveau:" + kansniveau,"Schadeniveau:" + schadeniveau, ncscdate,  productline,  ncscurl)
	

endrow = activeSheet.max_row + 1
activeSheet.delete_rows(2, endrow)

fontAlert = Font(color="FF0000")
fontNormal = Font(color="000000")

for regel in rssData:
    if regel in sheetData.keys():
        sheetData[regel]['Version'] = rssData[regel]['Version']
        if sheetData[regel]['Version'] != '1.00':
            sheetData[regel]['Vnew'] = 'yes'
        #activeSheet.append(sheetData[regel])
        #activeSheet['B' + str(regel)].font = fontAlert
    else:
        sheetData[regel] = rssData[regel]
        #activeSheet.append(sheetData[regel])

row = 2
#activeSheet.append(sheetData)
for rows in sheetData:
    activeSheet['A' + str(row)].value = rows
    activeSheet['B' + str(row)].value = sheetData[rows]['Version']
    if sheetData[rows]['Vnew'] == 'yes':
        activeSheet['B' + str(row)].font = fontAlert
    else:
        activeSheet['B' + str(row)].font = fontNormal
    activeSheet['C' + str(row)].value = sheetData[rows]['kansniveau']
    activeSheet['D' + str(row)].value = sheetData[rows]['schadeniveau']
    activeSheet['E' + str(row)].value = sheetData[rows]['publicatie datum']
    activeSheet['F' + str(row)].value = sheetData[rows]['product']
    activeSheet['G' + str(row)].value = sheetData[rows]['Beoordeling IB']
    activeSheet['H' + str(row)].value = sheetData[rows]['Beoordeeld']
    activeSheet['I' + str(row)].value = sheetData[rows]['Kwetsbaar']
    activeSheet['J' + str(row)].value = sheetData[rows]['uri']
    row = row + 1

#put everything in the excelsheet
try:
    wb.save(filename = ncsc_sheet_filename)
    print("All done: {}'s vulnerabilities added or updated!".format(str(row)))
    if online:
        print("Vulnerabilities were found online in the NCSC-feed {}".format(ncscurl))
    else:
        print("Vulnerabitities were imported locally from {}.".format(xmlimport_filename))
except:
    print("Something went wrong while saving the excelsheet")
